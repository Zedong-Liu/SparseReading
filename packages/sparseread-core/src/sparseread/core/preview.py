"""Deterministic no-HintSpec previews for Sparse Reading."""

from __future__ import annotations

import csv
import json
import re
import xml.etree.ElementTree as ET
import zipfile
from collections import Counter
from pathlib import Path
from typing import Any

import yaml

from sparseread.core.detector import FileInfo
from sparseread.core.models import CompressionInfo, FileCard, PreviewPack
from sparseread.core.readers.collection import CollectionReader
from sparseread.core.readers.text import TextReader


class PreviewBuilder:
    """Build L0 previews from file shape rather than task intent."""

    def __init__(self, collection_reader: CollectionReader, text_reader: TextReader) -> None:
        self.collection_reader = collection_reader
        self.text_reader = text_reader

    def build(self, info: FileInfo, card: FileCard, raw_ref: str) -> PreviewPack:
        try:
            if info.type == "collection":
                payload = self._collection(info.path)
            elif info.path.suffix.lower() in {".csv", ".tsv"}:
                payload = self._csv(info.path)
            elif info.path.suffix.lower() == ".xlsx":
                payload = self._xlsx(info.path)
            elif info.path.suffix.lower() == ".json":
                payload = self._mapping(info.path, "json")
            elif info.path.suffix.lower() in {".yaml", ".yml"}:
                payload = self._mapping(info.path, "yaml")
            elif info.path.suffix.lower() == ".xml":
                payload = self._xml(info.path)
            elif info.path.suffix.lower() == ".log":
                payload = self._log(info.path)
            else:
                payload = self._text(info.path)
        except Exception as exc:
            payload = {
                "recipe": "preview_error",
                "summary": f"preview error for {info.path.name}",
                "structure": {},
                "samples": [],
                "signals": [{"kind": "error", "message": str(exc)}],
                "visible_bytes": 0,
                "error": str(exc),
            }
        summary = str(payload.get("summary") or "")
        visible_bytes = int(payload.get("visible_bytes") or self._visible_size(payload))
        next_action = self._next_action(card)
        return PreviewPack(
            artifact_id=card.artifact_id,
            card=self._minimal_card(card),
            summary=summary,
            structure=dict(payload.get("structure") or {}),
            samples=list(payload.get("samples") or []),
            signals=list(payload.get("signals") or []),
            compression=CompressionInfo(
                recipe=str(payload.get("recipe") or "l0_default"),
                input_bytes=info.size_bytes,
                visible_bytes=visible_bytes,
                omitted=visible_bytes < info.size_bytes,
            ),
            raw_ref=raw_ref,
            next_action=next_action,
            error=str(payload.get("error") or ""),
        )

    @staticmethod
    def _minimal_card(card: FileCard) -> dict[str, Any]:
        allows_targeted_read = card.sparse_recommended or card.recommended_mode not in {"native", "native_read"}
        return {
            "path": card.path,
            "artifact_id": card.artifact_id,
            "type": card.type,
            "size_bytes": card.size_bytes,
            "structured": card.structured,
            "sparse_recommended": card.sparse_recommended,
            "reason": card.reason,
            "recommended_next": "sro_read_with_goal" if allows_targeted_read else "native",
            "recommended_mode": card.recommended_mode,
            "details": card.details,
        }

    @staticmethod
    def _next_action(card: FileCard) -> dict[str, Any]:
        allowed = ["use_preview", "sro_raw"]
        if card.sparse_recommended or card.recommended_mode not in {"native", "native_read"}:
            allowed.insert(1, "sro_read_with_goal")
        else:
            allowed.insert(1, "native")
        return {
            "allowed_next": allowed,
            "instruction": (
                "Use preview if sufficient. Call sro_read with artifact_id and a concrete "
                "HintSpec only for targeted evidence. Use sro_raw only when the original "
                "content is explicitly needed."
            ),
        }

    def _csv(self, path: Path) -> dict[str, Any]:
        delimiter = "\t" if path.suffix.lower() == ".tsv" else ","
        sampled_rows: list[dict[str, str]] = []
        row_count = 0
        with path.open(newline="", encoding="utf-8", errors="replace") as handle:
            reader = csv.DictReader(handle, delimiter=delimiter)
            headers = reader.fieldnames or []
            for row in reader:
                row_count += 1
                if len(sampled_rows) < 200:
                    sampled_rows.append({key: str(value or "") for key, value in row.items()})
        sample_rows = sampled_rows[:5]
        structure = {
            "columns": headers,
            "column_count": len(headers),
            "row_count": row_count,
            "column_types": self._column_types(headers, sampled_rows),
            "script_native_ok": True,
        }
        signals = self._tabular_signals(headers, sampled_rows)
        return {
            "recipe": "l0_csv_schema_sample_signals",
            "summary": f"{path.suffix.upper().lstrip('.')} with {row_count} rows and {len(headers)} columns",
            "structure": structure,
            "samples": sample_rows,
            "signals": signals,
        }

    def _xlsx(self, path: Path) -> dict[str, Any]:
        try:
            import openpyxl
        except ImportError:
            return self._xlsx_zip(path)
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheets = []
        samples: list[dict[str, Any]] = []
        try:
            for ws in wb.worksheets[:12]:
                rows_iter = ws.iter_rows(values_only=True)
                headers = [self._cell(value) for value in next(rows_iter, ())]
                first_rows = []
                for row in rows_iter:
                    values = [self._cell(value) for value in row]
                    if any(values):
                        first_rows.append({header: values[idx] if idx < len(values) else "" for idx, header in enumerate(headers)})
                    if len(first_rows) >= 5:
                        break
                sheets.append({
                    "name": ws.title,
                    "rows": ws.max_row,
                    "columns": ws.max_column,
                    "headers": headers,
                })
                samples.append({"sheet": ws.title, "rows": first_rows})
        finally:
            wb.close()
        return {
            "recipe": "l0_xlsx_sheet_schema_sample",
            "summary": f"XLSX workbook with {len(sheets)} sheets",
            "structure": {"sheets": sheets, "script_native_ok": True},
            "samples": samples[:4],
            "signals": [],
        }

    def _xlsx_zip(self, path: Path) -> dict[str, Any]:
        with zipfile.ZipFile(path) as zf:
            sheet_files = sorted(name for name in zf.namelist() if re.match(r"xl/worksheets/sheet\d+\.xml$", name))
        sheets = [{"name": f"Sheet{idx}", "source": name} for idx, name in enumerate(sheet_files, start=1)]
        return {
            "recipe": "l0_xlsx_zip_sheets",
            "summary": f"XLSX workbook with {len(sheets)} worksheet files",
            "structure": {"sheets": sheets, "script_native_ok": True},
            "samples": [],
            "signals": [],
        }

    def _mapping(self, path: Path, kind: str) -> dict[str, Any]:
        raw = path.read_text(encoding="utf-8", errors="replace")
        data = json.loads(raw) if kind == "json" else yaml.safe_load(raw)
        structure = self._schema_sketch(data)
        samples = self._mapping_samples(data)
        signals = self._mapping_signals(data)
        return {
            "recipe": f"l0_{kind}_schema_sample_signals",
            "summary": f"{kind.upper()} {structure.get('shape', 'value')} preview",
            "structure": structure,
            "samples": samples,
            "signals": signals,
        }

    def _xml(self, path: Path) -> dict[str, Any]:
        root = ET.parse(path).getroot()
        child_counts = Counter(child.tag for child in root)
        samples = []
        for elem in list(root.iter())[:12]:
            text = " ".join((elem.text or "").split())
            if text or elem.attrib:
                samples.append({"tag": elem.tag, "attrs": dict(elem.attrib), "text": self._clip(text, 180)})
        return {
            "recipe": "l0_xml_root_schema_sample",
            "summary": f"XML root <{root.tag}> with {sum(child_counts.values())} direct children",
            "structure": {"root": root.tag, "child_counts": dict(child_counts), "attributes": list(root.attrib)},
            "samples": samples[:8],
            "signals": [],
        }

    def _log(self, path: Path) -> dict[str, Any]:
        lines = path.read_text(encoding="utf-8", errors="replace").splitlines()
        severity_counts: Counter[str] = Counter()
        signature_counts: Counter[str] = Counter()
        notable: list[dict[str, Any]] = []
        for idx, line in enumerate(lines[:5000], start=1):
            severity = self._log_severity(line)
            if severity:
                severity_counts[severity] += 1
            signature = self._log_signature(line)
            if signature:
                signature_counts[signature] += 1
            if severity in {"ERROR", "WARN", "FATAL", "CRITICAL"} and len(notable) < 12:
                notable.append({"line": idx, "severity": severity, "text": self._clip(line, 260)})
        repeated = [
            {"signature": sig, "count": count}
            for sig, count in signature_counts.most_common(8)
            if count > 1
        ]
        return {
            "recipe": "l0_log_dedup_levels",
            "summary": f"log/text with {len(lines)} lines; levels={dict(severity_counts)}",
            "structure": {"line_count": len(lines), "level_counts": dict(severity_counts)},
            "samples": notable or [{"line": idx + 1, "text": self._clip(line, 180)} for idx, line in enumerate(lines[:5])],
            "signals": [{"kind": "repeated_lines", "items": repeated}] if repeated else [],
        }

    def _text(self, path: Path) -> dict[str, Any]:
        units, skeleton, kind = self.text_reader._load_units(path)
        samples = [{"anchor": unit.anchor, "text": self._clip(unit.text, 320)} for unit in units[:5]]
        signals = []
        headings = [line for line in skeleton if line.lower().startswith("heading")]
        if headings:
            signals.append({"kind": "headings", "items": headings[:12]})
        return {
            "recipe": f"l0_{kind}_skeleton_sample",
            "summary": f"{kind} object with {len(units)} text units",
            "structure": {"unit_count": len(units), "skeleton": skeleton[:24]},
            "samples": samples,
            "signals": signals,
        }

    def _collection(self, path: Path) -> dict[str, Any]:
        items = self.collection_reader._items(path)
        kind_counts = Counter(item.kind for item in items)
        notable = sorted(items, key=lambda item: item.size, reverse=True)[:12]
        return {
            "recipe": "l0_collection_grouped_card",
            "summary": f"collection with {len(items)} supported files",
            "structure": {
                "file_count": len(items),
                "kind_counts": dict(kind_counts),
                "total_bytes": sum(item.size for item in items),
            },
            "samples": [self.collection_reader._item_summary(item) for item in items[:12]],
            "signals": [
                {
                    "kind": "notable_files",
                    "items": [self.collection_reader._item_summary(item) for item in notable],
                }
            ],
        }

    @staticmethod
    def _column_types(headers: list[str], rows: list[dict[str, str]]) -> dict[str, str]:
        types: dict[str, str] = {}
        for header in headers:
            values = [row.get(header, "") for row in rows if row.get(header, "")]
            if not values:
                types[header] = "empty"
                continue
            if all(PreviewBuilder._looks_int(value) for value in values[:50]):
                types[header] = "integer"
            elif all(PreviewBuilder._looks_float(value) for value in values[:50]):
                types[header] = "number"
            elif all(PreviewBuilder._looks_date(value) for value in values[:50]):
                types[header] = "date"
            else:
                types[header] = "string"
        return types

    @staticmethod
    def _tabular_signals(headers: list[str], rows: list[dict[str, str]]) -> list[dict[str, Any]]:
        signals: list[dict[str, Any]] = []
        for header in headers:
            values = [row.get(header, "") for row in rows if row.get(header, "")]
            if not values:
                continue
            counts = Counter(values)
            rare = [value for value, count in counts.items() if count == 1 and PreviewBuilder._interesting_value(value)]
            if rare:
                signals.append({"kind": "rare_or_error_values", "column": header, "values": rare[:8]})
        return signals[:12]

    @classmethod
    def _mapping_signals(cls, data: Any) -> list[dict[str, Any]]:
        signals: list[dict[str, Any]] = []
        for path, value in cls._walk_scalars(data):
            text = str(value)
            if cls._interesting_value(text):
                signals.append({"kind": "interesting_scalar", "path": path, "value": cls._clip(text, 180)})
            if len(signals) >= 16:
                break
        return signals

    @classmethod
    def _mapping_samples(cls, data: Any) -> list[Any]:
        if isinstance(data, list):
            return data[:3]
        if isinstance(data, dict):
            return [{key: data[key] for key in list(data)[:8]}]
        return [data]

    @classmethod
    def _schema_sketch(cls, data: Any, depth: int = 0) -> dict[str, Any]:
        if depth > 4:
            return {"shape": type(data).__name__}
        if isinstance(data, dict):
            keys = list(data.keys())
            return {
                "shape": "object",
                "keys": keys[:24],
                "children": {str(key): cls._schema_sketch(data[key], depth + 1) for key in keys[:12]},
            }
        if isinstance(data, list):
            child = cls._schema_sketch(data[0], depth + 1) if data else {"shape": "empty"}
            return {"shape": "array", "length": len(data), "element": child}
        return {"shape": type(data).__name__}

    @classmethod
    def _walk_scalars(cls, data: Any, prefix: str = "$"):
        if isinstance(data, dict):
            for key, value in data.items():
                yield from cls._walk_scalars(value, f"{prefix}.{key}")
        elif isinstance(data, list):
            for idx, value in enumerate(data[:200]):
                yield from cls._walk_scalars(value, f"{prefix}[{idx}]")
        else:
            yield prefix, data

    @staticmethod
    def _visible_size(payload: dict[str, Any]) -> int:
        return len(json.dumps(payload, ensure_ascii=False, default=str))

    @staticmethod
    def _cell(value: Any) -> str:
        if value is None:
            return ""
        return str(value)

    @staticmethod
    def _clip(text: str, limit: int) -> str:
        text = text.strip()
        return text if len(text) <= limit else text[: limit - 15].rstrip() + "...[clipped]"

    @staticmethod
    def _interesting_value(value: str) -> bool:
        low = value.lower()
        return any(term in low for term in ("error", "warn", "fail", "exception", "timeout", "rate_limited", "critical", "panic"))

    @staticmethod
    def _looks_int(value: str) -> bool:
        return bool(re.fullmatch(r"[-+]?\d+", value.strip()))

    @staticmethod
    def _looks_float(value: str) -> bool:
        return bool(re.fullmatch(r"[-+]?(?:\d+\.\d+|\d+)", value.strip()))

    @staticmethod
    def _looks_date(value: str) -> bool:
        return bool(re.search(r"\b\d{4}[-/]\d{1,2}[-/]\d{1,2}\b", value))

    @staticmethod
    def _log_severity(line: str) -> str:
        match = re.search(r"\b(TRACE|DEBUG|INFO|WARN|WARNING|ERROR|CRITICAL|FATAL)\b", line, re.IGNORECASE)
        if not match:
            return ""
        severity = match.group(1).upper()
        return "WARN" if severity == "WARNING" else severity

    @classmethod
    def _log_signature(cls, line: str) -> str:
        text = re.sub(r"^\s*\d{4}[-/]\d{2}[-/]\d{2}[ T]\d{2}:\d{2}:\d{2}(?:[.,]\d+)?\s*", "", line)
        text = re.sub(r"^\s*\[[^\]]*\]\s*", "", text)
        text = re.sub(r"\b0x[0-9a-fA-F]+\b", "0x<hex>", text)
        text = re.sub(r"\b\d{4,}\b", "<num>", text)
        text = re.sub(r"\s+", " ", text).strip()
        return cls._clip(text, 180)

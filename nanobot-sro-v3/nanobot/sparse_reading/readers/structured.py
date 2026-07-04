"""Hint-guided reader for structured long files."""

from __future__ import annotations

import csv
import hashlib
import json
import re
import xml.etree.ElementTree as ET
import zipfile
from pathlib import Path
from typing import Any

import yaml

from nanobot.sparse_reading.models import EvidenceBlock, EvidencePack, HintSpec


class StructuredReader:
    """Extract schema, projections, samples, and exact field/key evidence."""

    _SMALL_TABLE_ROWS = 30

    def preview_details(self, path: Path, budget: int) -> dict[str, Any]:
        """Return a deterministic L0 preview without requiring a HintSpec."""
        suffix = path.suffix.lower()
        if suffix in {".csv", ".tsv"}:
            return self._preview_csv(path, budget, delimiter="\t" if suffix == ".tsv" else ",")
        if suffix == ".json":
            return self._preview_mapping(path, budget, "json")
        if suffix in {".yaml", ".yml"}:
            return self._preview_mapping(path, budget, "yaml")
        if suffix == ".xlsx":
            return self._preview_xlsx(path, budget)
        if suffix == ".xml":
            return self._preview_xml(path, budget)
        return {"kind": "structured_preview", "error": f"unsupported structured type: {path.suffix}"}

    def card_details(self, path: Path) -> dict[str, Any]:
        suffix = path.suffix.lower()
        if suffix in {".csv", ".tsv"}:
            delimiter = "\t" if suffix == ".tsv" else ","
            try:
                with path.open(newline="", encoding="utf-8", errors="replace") as f:
                    reader = csv.reader(f, delimiter=delimiter)
                    headers = next(reader, [])
                    row_count = sum(1 for _row in reader)
            except OSError:
                return {}
            return {
                "kind": "structured_card",
                "columns": headers,
                "column_count": len(headers),
                "row_count": row_count,
                "script_native_ok": True,
                "instruction": "For calculations or regressions, write a script that reads this local file path directly; do not request all rows into chat.",
            }
        if suffix == ".json":
            try:
                data = json.loads(path.read_text(encoding="utf-8", errors="replace"))
            except Exception:
                return {}
            keys = list(data.keys())[:24] if isinstance(data, dict) else []
            return {
                "kind": "structured_card",
                "top_keys": keys,
                "script_native_ok": True,
            }
        return {}

    def _preview_csv(self, path: Path, budget: int, *, delimiter: str) -> dict[str, Any]:
        try:
            with path.open(newline="", encoding="utf-8", errors="replace") as f:
                reader = csv.DictReader(f, delimiter=delimiter)
                headers = reader.fieldnames or []
                sample_rows: list[dict[str, str]] = []
                type_samples: dict[str, list[str]] = {header: [] for header in headers}
                row_count = 0
                for row in reader:
                    row_count += 1
                    if len(sample_rows) < 5:
                        sample_rows.append({header: row.get(header, "") for header in headers[:20]})
                    if row_count <= 100:
                        for header in headers:
                            value = str(row.get(header, "")).strip()
                            if value:
                                type_samples[header].append(value)
        except OSError as exc:
            return {"kind": "csv_preview", "error": str(exc)}
        preview = {
            "kind": "csv_preview",
            "row_count": row_count,
            "column_count": len(headers),
            "columns": headers[:40],
            "column_types": {
                header: self._infer_scalar_type(values)
                for header, values in type_samples.items()
            },
            "sample_rows": sample_rows,
            "default_read": "schema + first 5 rows; use a script on the local path for exact aggregation/full-table work",
        }
        return self._fit_preview(preview, budget)

    def _preview_mapping(self, path: Path, budget: int, kind: str) -> dict[str, Any]:
        try:
            raw = path.read_text(encoding="utf-8", errors="replace")
            data = json.loads(raw) if kind == "json" else yaml.safe_load(raw)
        except Exception as exc:
            return {"kind": f"{kind}_preview", "error": str(exc)}
        preview = {
            "kind": f"{kind}_preview",
            "shape": self._shape(data),
            "top_keys": list(data.keys())[:40] if isinstance(data, dict) else [],
            "path_index": self._mapping_path_index(data)[:40],
            "sample": self._sample_value(data),
            "default_read": "top-level structure + array/object schema; use sro_read only for targeted key/row evidence",
        }
        return self._fit_preview(preview, budget)

    def _preview_xlsx(self, path: Path, budget: int) -> dict[str, Any]:
        try:
            import openpyxl
        except ImportError:
            try:
                sheets = self._extract_xlsx_sheets_zip(path)
            except Exception as exc:
                return {"kind": "xlsx_preview", "error": str(exc)}
            preview = {
                "kind": "xlsx_preview",
                "backend": "stdlib_zip",
                "sheets": [
                    {
                        "name": name,
                        "row_count_sampled": max(0, len(rows) - 1),
                        "columns": rows[0][:40] if rows else [],
                        "sample_rows": rows[1:6] if len(rows) > 1 else [],
                    }
                    for name, rows in sheets[:8]
                ],
                "default_read": "sheet names + headers + first rows; use a script for exact workbook aggregation",
            }
            return self._fit_preview(preview, budget)
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            sheets: list[dict[str, Any]] = []
            for ws in wb.worksheets[:8]:
                rows_iter = ws.iter_rows(values_only=True)
                header = [self._cell(value) for value in next(rows_iter, ())]
                sample_rows: list[list[str]] = []
                for row in rows_iter:
                    if len(sample_rows) >= 5:
                        break
                    values = [self._cell(value) for value in row]
                    if any(values):
                        sample_rows.append(values[:20])
                sheets.append(
                    {
                        "name": ws.title,
                        "rows": ws.max_row,
                        "columns": ws.max_column,
                        "headers": header[:40],
                        "sample_rows": sample_rows,
                    }
                )
        finally:
            wb.close()
        return self._fit_preview(
            {
                "kind": "xlsx_preview",
                "sheets": sheets,
                "default_read": "sheet names + headers + first rows; use a script for exact workbook aggregation",
            },
            budget,
        )

    def _preview_xml(self, path: Path, budget: int) -> dict[str, Any]:
        try:
            root = ET.parse(path).getroot()
        except Exception as exc:
            return {"kind": "xml_preview", "error": str(exc)}
        preview = {
            "kind": "xml_preview",
            "root": root.tag,
            "path_index": self._xml_path_index(root)[:40],
            "sample_entries": [
                {"path": key, "value": self._short(value, 120)}
                for key, value in self._xml_entries(root)[:12]
            ],
            "default_read": "root/path index + first scalar entries; use sro_read for targeted elements",
        }
        return self._fit_preview(preview, budget)

    @classmethod
    def _fit_preview(cls, preview: dict[str, Any], budget: int) -> dict[str, Any]:
        """Trim verbose preview fields while keeping deterministic structure."""
        fitted = dict(preview)
        while len(json.dumps(fitted, ensure_ascii=False, default=str)) > budget:
            if fitted.get("sample_rows"):
                fitted["sample_rows"] = fitted["sample_rows"][:-1]
            elif isinstance(fitted.get("sample"), list) and fitted["sample"]:
                fitted["sample"] = fitted["sample"][:-1]
            elif fitted.get("path_index"):
                fitted["path_index"] = fitted["path_index"][: max(4, len(fitted["path_index"]) // 2)]
            elif fitted.get("columns"):
                fitted["columns"] = fitted["columns"][: max(8, len(fitted["columns"]) // 2)]
            else:
                fitted["truncated_to_budget"] = True
                break
        return fitted

    @staticmethod
    def _infer_scalar_type(values: list[str]) -> str:
        if not values:
            return "empty"
        sample = values[:50]
        if all(re.fullmatch(r"[-+]?\d+", value) for value in sample):
            return "int"
        if all(re.fullmatch(r"[-+]?(?:\d+\.\d*|\d*\.\d+|\d+)", value) for value in sample):
            return "number"
        lowered = {value.lower() for value in sample}
        if lowered <= {"true", "false", "yes", "no", "0", "1"}:
            return "bool"
        if all(re.fullmatch(r"\d{4}[-/]\d{2}[-/]\d{2}.*", value) for value in sample):
            return "date"
        return "string"

    @classmethod
    def _shape(cls, value: Any, depth: int = 0) -> Any:
        if depth >= 3:
            return type(value).__name__
        if isinstance(value, dict):
            return {
                "type": "object",
                "keys": list(value.keys())[:24],
                "fields": {str(key): cls._shape(item, depth + 1) for key, item in list(value.items())[:12]},
            }
        if isinstance(value, list):
            sample = next((item for item in value if item is not None), None)
            return {"type": "array", "count": len(value), "element": cls._shape(sample, depth + 1)}
        return type(value).__name__

    @classmethod
    def _sample_value(cls, value: Any, depth: int = 0) -> Any:
        if depth >= 2:
            return cls._short(value, 80)
        if isinstance(value, dict):
            return {str(key): cls._sample_value(item, depth + 1) for key, item in list(value.items())[:8]}
        if isinstance(value, list):
            return [cls._sample_value(item, depth + 1) for item in value[:5]]
        if value is None or isinstance(value, (bool, int, float, str)):
            return value
        return str(value)

    def read(self, path: Path, artifact_id: str, mode: str, hint: HintSpec, budget: int) -> EvidencePack:
        suffix = path.suffix.lower()
        if suffix in {".csv", ".tsv"}:
            return self._read_csv(path, artifact_id, mode, hint, budget, delimiter="\t" if suffix == ".tsv" else ",")
        if suffix == ".xlsx":
            return self._read_xlsx(path, artifact_id, mode, hint, budget)
        if suffix == ".json":
            return self._read_mapping(path, artifact_id, mode, hint, budget, "json")
        if suffix in {".yaml", ".yml"}:
            return self._read_mapping(path, artifact_id, mode, hint, budget, "yaml")
        if suffix == ".xml":
            return self._read_xml(path, artifact_id, mode, hint, budget)
        return EvidencePack(artifact_id=artifact_id, mode=mode, type="structured", summary="unsupported structured type", error=f"Unsupported structured file: {path}")

    def _read_csv(
        self, path: Path, artifact_id: str, mode: str, hint: HintSpec, budget: int, *, delimiter: str,
    ) -> EvidencePack:
        with path.open(newline="", encoding="utf-8", errors="replace") as f:
            reader = csv.DictReader(f, delimiter=delimiter)
            rows = list(reader)
            headers = reader.fieldnames or []
        skeleton = [
            f"columns({len(headers)}): {', '.join(headers[:30])}",
            f"rows: {len(rows)}",
        ]
        evidence: list[EvidenceBlock] = []
        calc_ready = None
        if self._should_return_full_table(mode, hint, len(rows)):
            selected_headers = self._project_headers(headers, hint)
            table = self._calc_ready_table(path, artifact_id, "rows", headers, rows, selected_headers)
            calc_ready = self._calc_ready_payload([table])
            evidence.append(EvidenceBlock("rows", self._calc_ready_anchor_text(table), 1.0))
            unresolved = []
            summary = self._subset_summary("CSV", [table])
        elif mode == "scout" or hint.want == "schema":
            sample = rows[:3]
            evidence.append(EvidenceBlock("schema", json.dumps({"columns": headers, "sample": sample}, ensure_ascii=False, indent=2), 1.0))
            evidence.extend(self._match_rows(rows, headers, hint, budget))
            unresolved = self._unresolved(hint, evidence)
            summary = f"CSV with {len(rows)} rows and {len(headers)} columns"
        else:
            evidence.extend(self._match_rows(rows, headers, hint, budget))
            unresolved = self._unresolved(hint, evidence)
            summary = f"CSV with {len(rows)} rows and {len(headers)} columns"
        return EvidencePack(
            artifact_id=artifact_id,
            mode=mode,
            type="csv",
            summary=summary,
            skeleton=skeleton,
            evidence=evidence[:10],
            unresolved=unresolved,
            calc_ready=calc_ready,
            next_action=self._calc_next_action(calc_ready) if calc_ready and not unresolved else None,
            next_hint=self._next_hint(hint, artifact_id) if unresolved and mode != "verify" else None,
        )

    def _read_xlsx(self, path: Path, artifact_id: str, mode: str, hint: HintSpec, budget: int) -> EvidencePack:
        try:
            import openpyxl
        except ImportError:
            return self._read_xlsx_zip(path, artifact_id, mode, hint, budget)
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        skeleton: list[str] = []
        evidence: list[EvidenceBlock] = []
        calc_tables: list[dict[str, Any]] = []
        terms = self._terms(hint)
        full_table_mode = False
        for ws in wb.worksheets:
            skeleton.append(f"sheet {ws.title}: {ws.max_row} rows x {ws.max_column} columns")
            if not self._sheet_relevant(ws.title, terms, mode):
                continue
            rows_iter = ws.iter_rows(values_only=True)
            header = [self._cell(v) for v in next(rows_iter, ())]
            selected_headers = self._project_headers(header, hint)
            wants_full_table = self._should_return_full_table(mode, hint, max(0, ws.max_row - 1))
            if header and not wants_full_table:
                evidence.append(EvidenceBlock(f"{ws.title}!header", ", ".join(header[:40]), 0.8))
            if wants_full_table:
                full_table_mode = True
                sheet_rows: list[dict[str, str]] = []
                for row in rows_iter:
                    values = [self._cell(v) for v in row]
                    if not any(values):
                        continue
                    if header:
                        sheet_rows.append(
                            {
                                header[idx]: values[idx] if idx < len(values) else ""
                                for idx in range(len(header))
                            }
                        )
                if header and sheet_rows:
                    table = self._calc_ready_table(
                        path, artifact_id, f"{ws.title}!rows", header, sheet_rows, selected_headers
                    )
                    calc_tables.append(table)
                    evidence.append(
                        EvidenceBlock(
                            f"{ws.title}!rows",
                            self._calc_ready_anchor_text(table),
                            1.0,
                        )
                    )
                continue
            if mode == "scout":
                for row_idx, row in enumerate(rows_iter, start=2):
                    values = [self._cell(v) for v in row]
                    joined = " | ".join(v for v in values if v)
                    if joined:
                        evidence.append(EvidenceBlock(f"{ws.title}!R{row_idx}", joined[:240], 0.3))
                    break
                continue
            for row_idx, row in enumerate(rows_iter, start=2):
                values = [self._cell(v) for v in row]
                joined = " | ".join(v for v in values if v)
                if not joined:
                    continue
                score = self._score_text(joined, terms)
                if score > 0 or (mode == "scout" and row_idx <= 4):
                    evidence.append(EvidenceBlock(f"{ws.title}!R{row_idx}", joined[:900], score or 0.1))
                if len(evidence) >= 24:
                    break
        wb.close()
        evidence.sort(key=lambda b: b.score, reverse=True)
        selected = self._fit_budget(evidence, budget)
        unresolved = [] if any(block.anchor.endswith("!rows") for block in selected) else self._unresolved(hint, selected)
        summary = f"XLSX with {len(skeleton)} sheets"
        calc_ready = self._calc_ready_payload(calc_tables) if calc_tables else None
        if full_table_mode and unresolved == []:
            summary = self._subset_summary("XLSX", calc_tables)
        return EvidencePack(
            artifact_id=artifact_id,
            mode=mode,
            type="xlsx",
            summary=summary,
            skeleton=skeleton[:12],
            evidence=selected,
            unresolved=unresolved,
            calc_ready=calc_ready,
            next_action=self._calc_next_action(calc_ready) if calc_ready and not unresolved else None,
            next_hint=self._next_hint(hint, artifact_id) if unresolved and mode != "verify" else None,
        )

    def _read_xlsx_zip(self, path: Path, artifact_id: str, mode: str, hint: HintSpec, budget: int) -> EvidencePack:
        """Small fallback for environments where openpyxl is unavailable."""
        try:
            sheets = self._extract_xlsx_sheets_zip(path)
        except Exception as exc:
            return EvidencePack(
                artifact_id=artifact_id,
                mode=mode,
                type="xlsx",
                summary="xlsx reader error",
                error=f"openpyxl unavailable and fallback XLSX extraction failed: {exc}",
                unresolved=list(hint.needles),
            )
        skeleton = [f"sheet {name}: {max(len(rows), 0)} rows" for name, rows in sheets]
        terms = self._terms(hint)
        evidence: list[EvidenceBlock] = []
        calc_tables: list[dict[str, Any]] = []
        full_table_mode = False
        for name, rows in sheets:
            if not self._sheet_relevant(name, terms, mode):
                continue
            header = rows[0] if rows else []
            selected_headers = self._project_headers(header, hint)
            wants_full_table = self._should_return_full_table(mode, hint, max(0, len(rows) - 1))
            if rows and not wants_full_table:
                evidence.append(EvidenceBlock(f"{name}!header", " | ".join(rows[0]), 0.8))
            if wants_full_table:
                full_table_mode = True
                sheet_rows: list[dict[str, str]] = []
                for row in rows[1:]:
                    if not any(row):
                        continue
                    sheet_rows.append(
                        {
                            header[idx]: row[idx] if idx < len(row) else ""
                            for idx in range(len(header))
                        }
                    )
                if header and sheet_rows:
                    table = self._calc_ready_table(
                        path, artifact_id, f"{name}!rows", header, sheet_rows, selected_headers
                    )
                    calc_tables.append(table)
                    evidence.append(
                        EvidenceBlock(
                            f"{name}!rows",
                            self._calc_ready_anchor_text(table),
                            1.0,
                        )
                    )
                continue
            if mode == "scout":
                if len(rows) > 1:
                    evidence.append(EvidenceBlock(f"{name}!R2", " | ".join(rows[1][:12]), 0.3))
                continue
            for idx, row in enumerate(rows[1:], start=2):
                text = " | ".join(cell for cell in row if cell)
                if not text:
                    continue
                score = self._score_text(text, terms)
                if score > 0 or (mode == "scout" and idx <= 4):
                    evidence.append(EvidenceBlock(f"{name}!R{idx}", text[:900], score or 0.1))
                if len(evidence) >= 24:
                    break
        evidence.sort(key=lambda b: b.score, reverse=True)
        selected = self._fit_budget(evidence, budget)
        unresolved = [] if any(block.anchor.endswith("!rows") for block in selected) else self._unresolved(hint, selected)
        summary = f"XLSX with {len(sheets)} sheets (stdlib fallback)"
        calc_ready = self._calc_ready_payload(calc_tables) if calc_tables else None
        if full_table_mode and unresolved == []:
            summary = self._subset_summary("XLSX", calc_tables)
        return EvidencePack(
            artifact_id=artifact_id,
            mode=mode,
            type="xlsx",
            summary=summary,
            skeleton=skeleton[:12],
            evidence=selected,
            unresolved=unresolved,
            calc_ready=calc_ready,
            next_action=self._calc_next_action(calc_ready) if calc_ready and not unresolved else None,
            next_hint=self._next_hint(hint, artifact_id) if unresolved and mode != "verify" else None,
        )

    def _extract_xlsx_sheets_zip(self, path: Path) -> list[tuple[str, list[list[str]]]]:
        ns = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
        with zipfile.ZipFile(path) as zf:
            shared = self._xlsx_shared_strings(zf, ns)
            names = self._xlsx_sheet_names(zf, ns)
            sheet_files = sorted(
                name for name in zf.namelist()
                if re.match(r"xl/worksheets/sheet\d+\.xml$", name)
            )
            sheets: list[tuple[str, list[list[str]]]] = []
            for idx, filename in enumerate(sheet_files, start=1):
                sheet_name = names[idx - 1] if idx - 1 < len(names) else f"Sheet{idx}"
                root = ET.fromstring(zf.read(filename))
                rows: list[list[str]] = []
                for row in root.findall(".//m:sheetData/m:row", ns):
                    values: list[str] = []
                    for cell in row.findall("m:c", ns):
                        values.append(self._xlsx_cell_value(cell, shared, ns))
                    if any(values):
                        rows.append(values)
                    if len(rows) >= 200:
                        break
                sheets.append((sheet_name, rows))
            return sheets

    @staticmethod
    def _xlsx_shared_strings(zf: zipfile.ZipFile, ns: dict[str, str]) -> list[str]:
        if "xl/sharedStrings.xml" not in zf.namelist():
            return []
        root = ET.fromstring(zf.read("xl/sharedStrings.xml"))
        out: list[str] = []
        for si in root.findall("m:si", ns):
            texts = [node.text or "" for node in si.findall(".//m:t", ns)]
            out.append("".join(texts))
        return out

    @staticmethod
    def _xlsx_sheet_names(zf: zipfile.ZipFile, ns: dict[str, str]) -> list[str]:
        if "xl/workbook.xml" not in zf.namelist():
            return []
        root = ET.fromstring(zf.read("xl/workbook.xml"))
        return [sheet.attrib.get("name", "") for sheet in root.findall(".//m:sheets/m:sheet", ns)]

    @staticmethod
    def _xlsx_cell_value(cell: ET.Element, shared: list[str], ns: dict[str, str]) -> str:
        if cell.attrib.get("t") == "inlineStr":
            return "".join(node.text or "" for node in cell.findall(".//m:t", ns))
        value = cell.find("m:v", ns)
        raw = value.text if value is not None else ""
        if cell.attrib.get("t") == "s":
            try:
                return shared[int(raw)]
            except Exception:
                return raw
        return raw or ""

    def _read_mapping(
        self, path: Path, artifact_id: str, mode: str, hint: HintSpec, budget: int, kind: str,
    ) -> EvidencePack:
        raw = path.read_text(encoding="utf-8", errors="replace")
        data = json.loads(raw) if kind == "json" else yaml.safe_load(raw)
        flat = list(self._flatten(data))
        path_index = self._mapping_path_index(data)
        skeleton = [*path_index[:16], *[f"{k}: {self._short(v, 120)}" for k, v in flat[:16]]]
        terms = self._terms(hint)
        evidence: list[EvidenceBlock] = []
        if mode == "scout" or hint.want == "schema":
            evidence.append(EvidenceBlock("path_index", "\n".join(path_index[:40]), 1.0))
        evidence.extend([
            EvidenceBlock(anchor=key, text=f"{key}: {self._short(value, 700)}", score=self._score_text(f"{key} {value}", terms))
            for key, value in flat
            if mode == "scout" or self._score_text(f"{key} {value}", terms) > 0
        ])
        evidence.sort(key=lambda b: b.score, reverse=True)
        selected = self._fit_budget(evidence[:30], budget)
        unresolved = self._unresolved(hint, selected)
        return EvidencePack(
            artifact_id=artifact_id,
            mode=mode,
            type=kind,
            summary=f"{kind.upper()} with {len(flat)} scalar/key-path entries",
            skeleton=skeleton,
            evidence=selected,
            unresolved=unresolved,
            next_hint=self._next_hint(hint, artifact_id) if unresolved and mode != "verify" else None,
        )

    def _read_xml(self, path: Path, artifact_id: str, mode: str, hint: HintSpec, budget: int) -> EvidencePack:
        root = ET.parse(path).getroot()
        entries = self._xml_entries(root)
        path_index = self._xml_path_index(root)
        terms = self._terms(hint)
        evidence: list[EvidenceBlock] = []
        if mode == "scout" or hint.want == "schema":
            evidence.append(EvidenceBlock("path_index", "\n".join(path_index[:40]), 1.0))
        evidence.extend(
            EvidenceBlock(anchor=tag, text=f"{tag}: {self._short(value, 700)}", score=self._score_text(f"{tag} {value}", terms))
            for tag, value in entries
            if mode == "scout" or self._score_text(f"{tag} {value}", terms) > 0
        )
        evidence.sort(key=lambda b: b.score, reverse=True)
        selected = self._fit_budget(evidence[:30], budget)
        unresolved = self._unresolved(hint, selected)
        return EvidencePack(
            artifact_id=artifact_id,
            mode=mode,
            type="xml",
            summary=f"XML root <{root.tag}> with {len(entries)} evidence-bearing elements sampled",
            skeleton=[f"root: {root.tag}", *path_index[:20]],
            evidence=selected,
            unresolved=unresolved,
            next_hint=self._next_hint(hint, artifact_id) if unresolved and mode != "verify" else None,
        )

    def _match_rows(self, rows: list[dict[str, str]], headers: list[str], hint: HintSpec, budget: int) -> list[EvidenceBlock]:
        terms = self._terms(hint)
        blocks: list[EvidenceBlock] = []
        for idx, row in enumerate(rows, start=2):
            text = " | ".join(f"{h}={row.get(h, '')}" for h in headers if row.get(h, ""))
            score = self._score_text(text, terms)
            if score > 0 or idx <= 4:
                blocks.append(EvidenceBlock(anchor=f"row {idx}", text=self._short(text, 900), score=score or 0.1))
            if len(blocks) >= 40:
                break
        blocks.sort(key=lambda b: b.score, reverse=True)
        return self._fit_budget(blocks, budget)

    def _compact_csv_rows(self, headers: list[str], rows: list[dict[str, str]]) -> str:
        compact_rows = [", ".join(headers)]
        compact_rows.extend(
            " | ".join(str(row.get(h, "")) for h in headers if row.get(h, "") != "")
            for row in rows
        )
        return "\n".join(compact_rows)[:2200]

    def _should_return_full_table(self, mode: str, hint: HintSpec, row_count: int) -> bool:
        if row_count <= 0 or row_count > self._SMALL_TABLE_ROWS:
            return False
        if hint.want not in {"table", "list"} and not self._requests_full_table(hint):
            return False
        if mode in {"focus", "refine"}:
            return True
        if mode == "scout":
            return self._requests_full_table(hint)
        return False

    @staticmethod
    def _requests_full_table(hint: HintSpec) -> bool:
        hay = " ".join([hint.goal, *hint.needles]).lower()
        phrases = (
            "all rows",
            "all data",
            "all records",
            "complete data",
            "complete records",
            "data rows",
            "remaining rows",
            "remaining records",
            "entire sheet",
            "full sheet",
            "all expense records",
        )
        if any(phrase in hay for phrase in phrases):
            return True
        return bool(
            re.search(r"\ball\s+\d+\s+[\w\s_-]{0,40}\b(?:rows|records)\b", hay)
        )

    def _flatten(self, obj: Any, prefix: str = "$"):
        if isinstance(obj, dict):
            for key, value in obj.items():
                yield from self._flatten(value, f"{prefix}.{key}")
        elif isinstance(obj, list):
            for idx, value in enumerate(obj[:200]):
                yield from self._flatten(value, f"{prefix}[{idx}]")
            if len(obj) > 200:
                yield f"{prefix}.length", str(len(obj))
        else:
            yield prefix, self._cell(obj)

    def _mapping_path_index(self, obj: Any, prefix: str = "$", depth: int = 0) -> list[str]:
        if depth > 6:
            return [f"{prefix}: ..."]
        if isinstance(obj, dict):
            keys = [str(key) for key in obj.keys()]
            out = [f"{prefix}: object keys={', '.join(keys[:18])}"]
            for key, value in list(obj.items())[:24]:
                out.extend(self._mapping_path_index(value, f"{prefix}.{key}", depth + 1))
                if len(out) >= 80:
                    break
            return out
        if isinstance(obj, list):
            item_types = sorted({type(item).__name__ for item in obj[:20]})
            out = [f"{prefix}: list len={len(obj)} item_types={', '.join(item_types)}"]
            if obj:
                out.extend(self._mapping_path_index(obj[0], f"{prefix}[]", depth + 1))
            return out
        return [f"{prefix}: {type(obj).__name__}"]

    def _xml_entries(self, root: ET.Element) -> list[tuple[str, str]]:
        entries: list[tuple[str, str]] = []

        def walk(elem: ET.Element, path: str) -> None:
            if len(entries) >= 300:
                return
            text = " ".join((elem.text or "").split())
            attrs = " ".join(f"@{k}={v}" for k, v in elem.attrib.items())
            value = " ".join(part for part in (attrs, text) if part)
            if value:
                entries.append((path, value))
            counts: dict[str, int] = {}
            for child in list(elem):
                counts[child.tag] = counts.get(child.tag, 0) + 1
                walk(child, f"{path}/{child.tag}[{counts[child.tag]}]")

        walk(root, f"/{root.tag}")
        return entries

    def _xml_path_index(self, root: ET.Element) -> list[str]:
        index: list[str] = []

        def walk(elem: ET.Element, path: str, depth: int) -> None:
            if depth > 6 or len(index) >= 80:
                return
            attrs = f" attrs={', '.join(elem.attrib.keys())}" if elem.attrib else ""
            children = list(elem)
            child_tags = list(dict.fromkeys(child.tag for child in children[:24]))
            child_text = f" children={', '.join(child_tags)}" if child_tags else ""
            index.append(f"{path}: element{attrs}{child_text}")
            counts: dict[str, int] = {}
            for child in children[:24]:
                counts[child.tag] = counts.get(child.tag, 0) + 1
                walk(child, f"{path}/{child.tag}[{counts[child.tag]}]", depth + 1)

        walk(root, f"/{root.tag}", 0)
        return index

    @staticmethod
    def _terms(hint: HintSpec) -> list[str]:
        goal_terms = re.findall(r"[A-Za-z0-9_./:-]{3,}", hint.goal)
        stop = {"the", "and", "with", "from", "that", "this", "find", "get", "all", "for", "sheet", "sheets"}
        terms: list[str] = []
        for term in [*hint.needles, *hint.must_keep, *goal_terms[:20]]:
            lowered = term.lower().strip()
            if lowered and lowered not in stop:
                terms.append(lowered)
        return terms

    @staticmethod
    def _score_text(text: str, terms: list[str]) -> float:
        hay = text.lower()
        score = 0.0
        for term in terms:
            if term in hay:
                score += 2.0
        if re.search(r"\b\d+(?:[.,]\d+)?\b", text):
            score += 0.5
        return score

    def _unresolved(self, hint: HintSpec, blocks: list[EvidenceBlock]) -> list[str]:
        text = "\n".join(block.text for block in blocks).lower()
        unresolved = [needle for needle in hint.needles if needle.lower() not in text]
        if self._requests_full_table(hint) and not any(
            block.anchor == "rows" or block.anchor.endswith("!rows") for block in blocks
        ):
            unresolved.append("complete rows")
        return unresolved

    def _fit_budget(self, blocks: list[EvidenceBlock], budget: int) -> list[EvidenceBlock]:
        picked: list[EvidenceBlock] = []
        seen: set[tuple[str, str]] = set()
        used = 0
        for block in blocks:
            dedupe_key = (block.anchor.split("!")[0], block.text)
            if dedupe_key in seen:
                continue
            cost = len(block.anchor) + len(block.text) + 32
            if picked and used + cost > budget:
                continue
            picked.append(block)
            seen.add(dedupe_key)
            used += cost
            if used >= budget:
                break
        return picked[:8]

    @staticmethod
    def _sheet_relevant(sheet_name: str, terms: list[str], mode: str) -> bool:
        if mode == "scout" or not terms:
            return True
        name = sheet_name.lower()
        generic = {"all", "row", "rows", "record", "records", "table", "tables", "complete", "sheet", "data"}
        tokens = {
            token
            for term in terms
            for token in re.findall(r"[A-Za-z0-9_./:-]{3,}", term.lower())
            if token not in generic
        }
        if not tokens:
            return True
        return any(token in name for token in tokens)

    @staticmethod
    def _project_headers(headers: list[str], hint: HintSpec) -> list[str]:
        if not headers:
            return []
        terms = StructuredReader._terms(hint)
        if not terms:
            return headers
        selected = [header for header in headers if StructuredReader._header_matches(header, terms)]
        if not selected:
            return headers
        if len(selected) < 2:
            for header in headers:
                if header in selected:
                    continue
                if StructuredReader._is_context_header(header):
                    selected.append(header)
                    if len(selected) >= 2:
                        break
        selected_set = set(selected)
        return [header for header in headers if header in selected_set]

    @staticmethod
    def _header_matches(header: str, terms: list[str]) -> bool:
        compact_header = re.sub(r"[^a-z0-9]+", "", header.lower())
        if not compact_header:
            return False
        for term in terms:
            compact_term = re.sub(r"[^a-z0-9]+", "", term.lower())
            if compact_term and (compact_term in compact_header or compact_header in compact_term):
                return True
        return False

    @staticmethod
    def _is_context_header(header: str) -> bool:
        return re.sub(r"[^a-z0-9]+", "", header.lower()) in {
            "id",
            "name",
            "date",
            "time",
            "region",
            "category",
            "type",
            "product",
            "item",
            "label",
            "title",
            "amount",
            "value",
            "count",
            "total",
            "revenue",
            "cost",
            "price",
            "employee",
            "customer",
            "account",
            "status",
        }

    @staticmethod
    def _subset_summary(kind: str, tables: list[dict[str, Any]]) -> str:
        if not tables:
            return f"{kind} subset ready"
        parts: list[str] = []
        for table in tables:
            name = str(table.get("name") or "").strip()
            row_count = table.get("row_count", 0)
            column_count = table.get("column_count", 0)
            label = f"{name}: " if name and name != "rows" else ""
            parts.append(f"{label}{row_count} rows x {column_count} columns")
        return f"{kind} subset ready: {'; '.join(parts)}"

    @staticmethod
    def _calc_ready_anchor_text(table: dict[str, Any]) -> str:
        columns = [str(column) for column in table.get("columns", []) if str(column)]
        row_count = table.get("row_count", 0)
        column_count = table.get("column_count", len(columns))
        name = str(table.get("name") or "rows").strip() or "rows"
        column_text = ", ".join(columns[:24])
        return f"{name}: {row_count} rows x {column_count} columns; columns: {column_text}"

    def _calc_ready_table(
        self,
        source_path: Path,
        artifact_id: str,
        anchor: str,
        headers: list[str],
        rows: list[dict[str, str]],
        selected_headers: list[str] | None = None,
    ) -> dict[str, Any]:
        columns = selected_headers or headers
        calc_path = self._calc_table_path(source_path, artifact_id, anchor, headers, columns)
        name = anchor[:-5] if anchor.endswith("!rows") else anchor
        lines = ["	".join(columns)]
        lines.extend(
            "	".join(str(row.get(header, "")) for header in columns)
            for row in rows
        )
        calc_path.parent.mkdir(parents=True, exist_ok=True)
        calc_path.write_text("\n".join(lines), encoding="utf-8")
        return {
            "anchor": anchor,
            "name": name,
            "row_count": len(rows),
            "column_count": len(columns),
            "columns": columns,
            "tsv_path": str(calc_path),
        }

    @staticmethod
    def _calc_ready_payload(tables: list[dict[str, Any]]) -> dict[str, Any]:
        payload = {
            "kind": "structured_rows",
            "instructions": "Use the TSV artifact(s) in calc_ready['tables'] as the exact source for the next calculation step.",
            "tables": tables,
        }
        payload["python_variable"] = "calc_ready"
        payload["python_prelude"] = (
            "import csv; tables = {t['name']: list(csv.DictReader(open(t['tsv_path'], newline='', encoding='utf-8'), delimiter='	')) "
            "for t in calc_ready['tables']}"
        )
        return payload

    @staticmethod
    def _calc_next_action(calc_ready: dict[str, Any]) -> dict[str, Any]:
        return {
            "tool": "exec",
            "priority": "immediate",
            "reason": "Exact structured subset is already materialized.",
            "instructions": [
                "Use calc_ready['tables'][*]['tsv_path'] in one short calculation script.",
                "Do not reread the source object.",
            ],
        }

    @staticmethod
    def _calc_table_path(source_path: Path, artifact_id: str, anchor: str, headers: list[str], columns: list[str]) -> Path:
        safe_anchor = re.sub(r"[^A-Za-z0-9_.-]+", "_", anchor).strip("._") or "rows"
        if columns and columns != headers:
            col_sig = hashlib.sha1("\0".join(columns).encode("utf-8")).hexdigest()[:8]
            safe_anchor = f"{safe_anchor}_{col_sig}"
        return source_path.parent / ".nanobot" / "sro-calc" / artifact_id / f"{safe_anchor}.tsv"

    @staticmethod
    def _next_hint(hint: HintSpec, artifact_id: str) -> dict[str, Any]:
        return {
            "goal": hint.goal,
            "needles": hint.needles[:6],
            "want": hint.want,
            "scope": "narrow",
            "artifact": artifact_id,
            "type_hint": hint.type_hint,
            "must_keep": hint.must_keep,
        }

    @staticmethod
    def _cell(value: Any) -> str:
        return "" if value is None else str(value)

    @staticmethod
    def _short(value: Any, limit: int) -> str:
        text = str(value)
        return text if len(text) <= limit else text[: limit - 18].rstrip() + "...[clipped]"
